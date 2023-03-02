from random import randint
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.append(["번호","영어","수학"])
for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100)])

# 전체 rows
print(tuple(ws.rows)) # 가로 한 줄씩 가지고 오기

# 전체 columns 
print(tuple(ws.columns)) # 세로 한 줄씩 가지고오기

for row in tuple(ws.rows):
    print(row[1].value)

for column in tuple(ws.columns):
    print(column[1].value)

for rows in ws.iter_rows(): # 전체 row
    print(row[1].value)

for column in ws.iter_cols(): # ()값을 넣어서 범위 지정 가능
    print(column[1].value)

wb.save("sample.xlsx")