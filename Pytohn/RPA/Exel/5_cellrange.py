from random import randint
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 한줄씩 데이터 넣기
ws.append(["번호","영어","수학"])
for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100)])

col_b = ws["B"] # 영어 column만 가지고 오기
print(col_b)

for cell in col_b:
    print(cell.value)

col_range = ws["B:C"] # 영어, 수학 column 함께 가지고오기
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1] # 1번째 row만 가지고오기
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가지고오기
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

from openpyxl.utils.cell import coordinate_from_string
row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ") # 데이터 셀의 위치 정보 표시
        xy = coordinate_from_string(cell.coordinate) # 데이터 셀의 위치정보를 튜플형태로 나누어줌
        # print(xy, end=" ")
        print(xy[0], end="") # A
        print(xy[1], end=" ") # 1
    print()

wb.save("sample.xlsx")