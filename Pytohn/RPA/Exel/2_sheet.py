from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet를 기본 이름으로 생성
ws.title = "MySheet" # sheet의 이름 변경
ws.sheet_properties.tabColor = "ff66ff" #RGB형태로 값을 넣어주면 텝 색상 변경

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("NewSheet",2) #2번째 index에 sheet 생성
 
new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근

print(wb.sheetnames) # 모든 Sheet내용 확인

# Sheet 복사
new_ws["A1"] ="Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")