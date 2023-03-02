from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "MySheet"

# A1 셀에 1이라는 값을 입력
ws["A1"] =1
ws["A2"] =2
ws["A3"] =3

ws["B1"] =4
ws["B2"] =5
ws["B3"] =6

print(ws["A1"]) # A1셀의 정보를 출력
print(ws["A1"].value) # A1 셀의 값을 출력
print(ws["A10"].value) # A10의 값을 출력 / 값이 없으면 None을 출력

print(ws.cell(row=1,column=1).value) # A1의 값
print(ws.cell(row=1,column=2).value) # B1의 값

c = ws.cell(column=3, row=1, value=10) # C1의 값을 10으로 설정 
print(c.value)

from random import *
for x in range(1,11):
    for y in range(1,11):
        ws.cell(row=x, column=y, value=randint(0,100))


wb.save("sample.xlsx")